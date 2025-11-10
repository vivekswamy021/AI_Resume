import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
import traceback
import re
from dotenv import load_dotenv 
from datetime import date 
import csv 
from streamlit.runtime.uploaded_file_manager import UploadedFile

# Import the dashboard files (will be created separately)
from admin_dashboard import admin_dashboard
from candidate_dashboard import candidate_dashboard
from hiring_dashboard import hiring_dashboard

# =========================================================
# CORE APPLICATION SETUP, UTILITIES, AND LLM FUNCTIONS
# =========================================================

# -------------------------
# CONFIGURATION & API SETUP
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"

# Options for LLM functions
section_options = ["name", "email", "phone", "skills", "education", "experience", "certifications", "projects", "strength", "personal_details", "github", "linkedin", "full resume"]
question_section_options = ["skills","experience", "certifications", "projects", "education"] 

# Default Categories for JD Filtering
DEFAULT_JOB_TYPES = ["Full-time", "Contract", "Internship", "Remote", "Part-time"]
DEFAULT_ROLES = ["Software Engineer", "Data Scientist", "Product Manager", "HR Manager", "Marketing Specialist", "Operations Analyst"]

# Load environment variables from .env file
load_dotenv()

# Ensure GROQ_API_KEY is defined
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

if not GROQ_API_KEY:
    st.warning(
        "🚨 WARNING: GROQ_API_KEY environment variable not set. "
        "AI functionality (Parsing, Matching, Q&A) will not work. "
        "Please ensure a '.env' file exists with your key."
    )
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set. AI functions disabled.")
            return Completions()
    
    client = MockGroqClient()
else:
    # Initialize Groq Client
    client = Groq(api_key=GROQ_API_KEY)


# -------------------------
# Utility: Navigation Manager
# -------------------------
def go_to(page_name):
    """Changes the current page in Streamlit's session state."""
    st.session_state.page = page_name

def clear_interview_state():
    """Clears all generated questions, answers, and the evaluation report."""
    st.session_state.interview_qa = []
    st.session_state.iq_output = ""
    st.session_state.evaluation_report = ""
    st.toast("Practice answers cleared.")

# -------------------------
# CORE LOGIC: FILE HANDLING AND EXTRACTION
# -------------------------

def get_file_type(file_path):
    """Identifies the file type based on its extension."""
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    
    if ext == 'pdf':
        return 'pdf'
    elif ext == 'docx':
        return 'docx'
    elif ext == 'xlsx':
        return 'xlsx'
    elif ext in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf']:
        return ext
    else:
        return 'txt' 

def extract_content(file_type, file_path):
    """Extracts text content from various file types."""
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
        
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                text += f"--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        text += row_text + '\n'
                text += "\n"
        
        elif file_type == 'csv':
             with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += ' | '.join(row) + '\n'
        
        # Handles txt, json, md, markdown, rtf, and the default case
        elif file_type in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf'] or file_type not in ['pdf', 'docx', 'xlsx']:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

        if not text.strip():
            return f"Error: {file_type.upper()} content extraction failed. The file appears empty or non-text content could not be read."
        
        return text
    
    except Exception as e:
        return f"Fatal Extraction Error: Failed to read file content ({file_type}). Error details: {e}"

# -------------------------
# LLM & Extraction Functions
# -------------------------

@st.cache_data(show_spinner="Extracting JD metadata...")
def extract_jd_metadata(jd_text):
    """Extracts structured metadata (Role, Job Type, Key Skills) from raw JD text."""
    if not GROQ_API_KEY:
        return {"role": "N/A", "job_type": "N/A", "key_skills": []}

    prompt = f"""Analyze the following Job Description and extract the key metadata.
    
    Job Description:
    {jd_text}
    
    Provide the output strictly as a JSON object with the following three keys:
    1.  **role**: The main job title (e.g., 'Data Scientist', 'Senior Software Engineer'). If not clear, default to 'General Analyst'.
    2.  **job_type**: The employment type (e.g., 'Full-time', 'Contract', 'Internship', 'Remote'). If not clear, default to 'Full-time'.
    3.  **key_skills**: A list of 5 to 10 most critical hard and soft skills required (e.g., ['Python', 'AWS', 'Teamwork', 'SQL']).
    
    Example Output: {{"role": "Software Engineer", "job_type": "Full-time", "key_skills": ["Python", "JavaScript", "React", "AWS", "Agile"]}}
    """
    content = ""
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        content = response.choices[0].message.content.strip()

        # Robust JSON extraction
        json_str = content
        if json_str.startswith('```json'):
            json_str = json_str[len('```json'):]
        if json_str.endswith('```'):
            json_str = json_str[:-len('```')]
        json_str = json_str.strip()

        json_start = json_str.find('{')
        json_end = json_str.rfind('}') + 1

        if json_start != -1 and json_end != -1 and json_end > json_start:
            json_str = json_str[json_start:json_end]

        parsed = json.loads(json_str)
        
        return {
            "role": parsed.get("role", "General Analyst"),
            "job_type": parsed.get("job_type", "Full-time"),
            "key_skills": [s.strip() for s in parsed.get("key_skills", []) if isinstance(s, str)]
        }

    except Exception:
        return {"role": "General Analyst (LLM Error)", "job_type": "Full-time (LLM Error)", "key_skills": ["LLM Error", "Fallback"]}


@st.cache_data(show_spinner="Analyzing content with Groq LLM...")
def parse_with_llm(text, return_type='json'):
    """Sends resume text to the LLM for structured information extraction. 
    
    Includes the fix for 'Extra data' JSON decoding errors."""
    if text.startswith("Error"):
        return {"error": text, "raw_output": ""}
    if not GROQ_API_KEY:
        return {"error": "GROQ_API_KEY not set. Cannot run LLM parsing.", "raw_output": ""}

    prompt = f"""Extract the following information from the resume in structured JSON.
    Ensure all relevant details for each category are captured.
    - Name, - Email, - Phone, - Skills, - Education (list of degrees/institutions/dates), 
    - Experience (list of job roles/companies/dates/responsibilities), - Certifications (list), 
    - Projects (list of project names/descriptions/technologies), - Strength (list of personal strengths/qualities), 
    - Personal Details (e.g., address, date of birth, nationality), - Github (URL), - LinkedIn (URL)
    
    Resume Text:
    {text}
    
    Provide the output strictly as a JSON object.
    """
    content = ""
    parsed = {}
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
        content = response.choices[0].message.content.strip()

        # --- FIX: AGGRESSIVELY ISOLATE THE JSON OBJECT ---
        json_str = content
        
        # 1. Strip common LLM fences and surrounding whitespace
        if json_str.startswith('```json'):
            json_str = json_str[len('```json'):]
        if json_str.endswith('```'):
            json_str = json_str[:-len('```')]
        json_str = json_str.strip()

        # 2. Find the index of the first '{' and the last '}'
        json_start = json_str.find('{')
        json_end = json_str.rfind('}') + 1 

        # 3. CRITICAL: Only slice the content between the first '{' and the last '}'
        if json_start != -1 and json_end != -1 and json_end > json_start:
            json_str = json_str[json_start:json_end]
        else:
            raise json.JSONDecodeError("Could not isolate a valid JSON structure from LLM response.", content, 0)

        parsed = json.loads(json_str)

    except json.JSONDecodeError as e:
        error_msg = f"JSON decoding error from LLM. LLM returned malformed JSON. Error: {e}"
        parsed = {"error": error_msg, "raw_output": content}
    except ValueError as e:
        parsed = {"error": str(e), "raw_output": "AI functions disabled."}
    except Exception as e:
        error_msg = f"LLM API interaction error: {e}"
        parsed = {"error": error_msg, "raw_output": "No LLM response due to API error."}

    if return_type == 'json':
        return parsed
    elif return_type == 'markdown':
        if "error" in parsed:
            return f"**Error:** {parsed.get('error', 'Unknown parsing error')}\nRaw output:\n```\n{parsed.get('raw_output','')}\n```"
        
        md = ""
        for k, v in parsed.items():
            if v:
                md += f"**{k.replace('_', ' ').title()}**:\n"
                if isinstance(v, list):
                    for item in v:
                        if item: 
                            md += f"- {item}\n"
                elif isinstance(v, dict):
                    for sub_k, sub_v in v.items():
                        if sub_v:
                            md += f"  - {sub_k.replace('_', ' ').title()}: {sub_v}\n"
                    md += "\n"
                else:
                    md += f"  {v}\n"
                md += "\n"
        return md
    return {"error": "Invalid return_type"}


def extract_jd_from_linkedin_url(url: str) -> str:
    """
    Simulates JD content extraction from a LinkedIn URL.
    """
    try:
        job_title = "Data Scientist"
        try:
            match = re.search(r'/jobs/view/([^/]+)', url) or re.search(r'/jobs/(\w+)', url)
            if match:
                job_title = match.group(1).split('?')[0].replace('-', ' ').title()
                if job_title.lower().startswith('view'): job_title = 'Data Scientist'
        except:
            pass

        if "linkedin.com/jobs/" not in url:
             return f"[Error: Not a valid LinkedIn Job URL format: {url}]"

        jd_text = f"""
        --- Simulated JD for: {job_title} ---
        
        **Company:** Quantum Analytics Inc.
        **Role:** {job_title}
        
        **Responsibilities:**
        - Develop and implement machine learning models to solve complex business problems.
        - Clean, transform, and analyze large datasets using Python/R and SQL.
        - Collaborate with engineering teams to deploy models into production environments.
        - Communicate findings and model performance to non-technical stakeholders.
        
        **Requirements:**
        - MS/PhD in Computer Science, Statistics, or a quantitative field.
        - 3+ years of experience as a Data Scientist.
        - Expertise in Python (Pandas, Scikit-learn, TensorFlow/PyTorch).
        - Experience with cloud platforms (AWS, Azure, or GCP).
        
        --- End Simulated JD ---
        """
        
        return jd_text.strip()
            
    except Exception as e:
        return f"[Fatal Extraction Error: Simulation failed for URL {url}. Error: {e}]"


def evaluate_jd_fit(job_description, parsed_json):
    """Evaluates how well a resume fits a given job description."""
    if not GROQ_API_KEY:
        return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if not job_description.strip(): return "Please paste a job description."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."
    
    relevant_resume_data = {
        'Skills': parsed_json.get('skills', 'Not found or empty'),
        'Experience': parsed_json.get('experience', 'Not found or empty'),
        'Education': parsed_json.get('education', 'Not found or empty'),
    }
    resume_summary = json.dumps(relevant_resume_data, indent=2)

    prompt = f"""Evaluate how well the following resume content matches the provided job description.
    
    Job Description: {job_description}
    
    Resume Sections for Analysis:
    {resume_summary}
    
    Provide a detailed evaluation structured as follows:
    1.  **Overall Fit Score:** A score out of 10.
    2.  **Section Match Percentages:** A percentage score for the match in the key sections (Skills, Experience, Education).
    3.  **Strengths/Matches:** Key points where the resume aligns well with the JD.
    4.  **Gaps/Areas for Improvement:** Key requirements in the JD that are missing or weak in the resume.
    5.  **Overall Summary:** A concise summary of the fit.
    
    **Format the output strictly as follows, ensuring the scores are easily parsable (use brackets or no brackets around scores):**
    Overall Fit Score: [Score]/10
    
    --- Section Match Analysis ---
    Skills Match: [XX]%
    Experience Match: [YY]%
    Education Match: [ZZ]%
    
    Strengths/Matches:
    - Point 1
    - Point 2
    
    Gaps/Areas for Improvement:
    - Point 1
    - Point 2
    
    Overall Summary: [Concise summary]
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def evaluate_interview_answers(qa_list, parsed_json):
    """Evaluates the user's answers against the resume content and provides feedback."""
    if not GROQ_API_KEY:
        return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."

    resume_summary = json.dumps(parsed_json, indent=2)
    
    qa_summary = "\n---\n".join([
        f"Q: {item['question']}\nA: {item['answer']}" 
        for item in qa_list
    ])
    
    prompt = f"""You are an expert HR Interviewer. Evaluate the candidate's answers based on the following:
    1.  **The Candidate's Resume Content (for context):**
        {resume_summary}
    2.  **The Candidate's Questions and Answers:**
        {qa_summary}

    For each Question-Answer pair, provide a score (out of 10) and detailed feedback. The feedback must include:
    * **Clarity & Accuracy:** How well the answer directly and accurately addresses the question, referencing the resume context.
    * **Gaps & Improvements:** Specific suggestions on how the candidate could improve the answer or what critical resume points they missed/could elaborate on.
    
    Finally, provide an **Overall Summary** and a **Total Score (out of {len(qa_list) * 10})**.
    
    **Format the output strictly using Markdown headings and bullet points:**
    
    ## Evaluation Results
    
    ### Question 1: [Question Text]
    Score: [X]/10
    Feedback:
    - **Clarity & Accuracy:** ...
    - **Gaps & Improvements:** ...
    
    ... [Repeat for all questions] ...
    
    ---
    
    ## Final Assessment
    Total Score: [Y]/{len(qa_list) * 10}
    Overall Summary: [A concise summary of the candidate's performance and next steps.]
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def generate_interview_questions(parsed_json, section):
    """Generates categorized interview questions using LLM."""
    if not GROQ_API_KEY:
        return "AI Functions Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot generate questions due to resume parsing errors."
    
    section_title = section.replace("_", " ").title()
    section_content = parsed_json.get(section, "")
    if isinstance(section_content, (list, dict)):
        section_content = json.dumps(section_content, indent=2)
    elif not isinstance(section_content, str):
        section_content = str(section_content)

    if not section_content.strip():
        return f"No significant content found for the '{section_title}' section in the parsed resume. Please select a section with relevant data to generate questions."

    prompt = f"""Based on the following {section_title} section from the resume: {section_content}
Generate 3 interview questions each for these levels: Generic, Basic, Intermediate, Difficult.
**IMPORTANT: Format the output strictly as follows, with level headers and questions starting with 'Qx:':**
[Generic]
Q1: Question text...
Q2: Question text...
Q3: Question text...
[Basic]
Q1: Question text...
...
[Difficult]
Q3: Question text...
    """
    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.5
    )
    return response.choices[0].message.content.strip()


# -------------------------
# Utility Functions
# -------------------------
def dump_to_excel(parsed_json, filename):
    """Dumps parsed JSON data to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Data"
    ws.append(["Category", "Details"])
    
    section_order = ['name', 'email', 'phone', 'github', 'linkedin', 'experience', 'education', 'skills', 'projects', 'certifications', 'strength', 'personal_details']
    
    for section_key in section_order:
        if section_key in parsed_json and parsed_json[section_key]:
            content = parsed_json[section_key]
            
            if section_key in ['name', 'email', 'phone', 'github', 'linkedin']:
                ws.append([section_key.replace('_', ' ').title(), str(content)])
            else:
                ws.append([])
                ws.append([section_key.replace('_', ' ').title()])
                
                if isinstance(content, list):
                    for item in content:
                        if item:
                            ws.append(["", str(item)])
                elif isinstance(content, dict):
                    for k, v in content.items():
                        if v:
                            ws.append(["", f"{k.replace('_', ' ').title()}: {v}"])
                else:
                    ws.append(["", str(content)])

    wb.save(filename)
    with open(filename, "rb") as f:
        return f.read()

def parse_and_store_resume(file_input, file_name_key='default', source_type='file'):
    """
    Handles file/text input, parsing, and stores results.
    """
    
    text = None
    file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"

    if source_type == 'file':
        if not isinstance(file_input, UploadedFile):
            return {"error": "Invalid file input type passed to parser.", "full_text": ""}

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file_input.name) 
        with open(temp_path, "wb") as f:
            f.write(file_input.getbuffer()) 

        file_type = get_file_type(temp_path)
        text = extract_content(file_type, temp_path)
        file_name = file_input.name.split('.')[0]
    
    elif source_type == 'text':
        text = file_input
        file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"
        
    if text.startswith("Error"):
        return {"error": text, "full_text": text, "name": file_name}

    parsed = parse_with_llm(text, return_type='json')
    
    if not parsed or "error" in parsed:
        return {"error": parsed.get('error', 'Unknown parsing error'), "full_text": text, "name": file_name}

    excel_data = None
    if file_name_key == 'single_resume_candidate':
        try:
            name = parsed.get('name', 'candidate').replace(' ', '_').strip()
            name = "".join(c for c in name if c.isalnum() or c in ('_', '-')).rstrip()
            if not name: name = "candidate"
            excel_filename = os.path.join(tempfile.gettempdir(), f"{name}_parsed_data.xlsx")
            excel_data = dump_to_excel(parsed, excel_filename)
        except Exception as e:
            pass
    
    final_name = parsed.get('name', file_name)

    return {
        "parsed": parsed,
        "full_text": text,
        "excel_data": excel_data,
        "name": final_name
    }


def qa_on_resume(question):
    """Chatbot for Resume (Q&A) using LLM."""
    if not GROQ_API_KEY:
        return "AI Chatbot Disabled: GROQ_API_KEY not set."
        
    parsed_json = st.session_state.parsed
    full_text = st.session_state.full_text
    prompt = f"""Given the following resume information:
    Resume Text: {full_text}
    Parsed Resume Data (JSON): {json.dumps(parsed_json, indent=2)}
    Answer the following question about the resume concisely and directly.
    If the information is not present, state that clearly.
    Question: {question}
    """
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.4)
    return response.choices[0].message.content.strip()

# -------------------------
# UI PAGES: Authentication (Login, Signup)
# -------------------------

def login_page():
    st.title("🌐 PragyanAI Job Portal")
    st.header("Login")

    selected_role = st.selectbox(
        "Select Your Role",
        ["Select Role", "Admin Dashboard", "Candidate Dashboard", "Hiring Company Dashboard"],
        key="login_role_select"
    )
    
    st.markdown("---")

    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    if st.button("Login", use_container_width=True):
        if email and password:
            if selected_role == "Select Role":
                st.error("Please select your role before logging in.")
            elif selected_role == "Admin Dashboard":
                st.success("Login successful! Redirecting to Admin Dashboard.")
                go_to("admin_dashboard")
            elif selected_role == "Candidate Dashboard":
                st.success("Login successful! Redirecting to Candidate Dashboard.")
                go_to("candidate_dashboard")
            elif selected_role == "Hiring Company Dashboard":
                st.success("Login successful! Redirecting to Hiring Company Dashboard.")
                go_to("hiring_dashboard")
        else:
            st.error("Please enter both email and password")

    st.markdown("---")
    
    if st.button("Don't have an account? Sign up here"):
        go_to("signup")

def signup_page():
    st.header("Create an Account")
    email = st.text_input("Email")
    password = st.text_input("Password", type="password")
    confirm = st.text_input("Confirm Password", type="password")

    if st.button("Sign Up", use_container_width=True):
        if password == confirm and email:
            st.success("Signup successful! Please login.")
            go_to("login")
        else:
            st.error("Passwords do not match or email is empty")

    if st.button("Already have an account? Login here"):
        go_to("login")

# =========================================================
# MAIN APP EXECUTION BLOCK
# =========================================================

def main():
    st.set_page_config(layout="wide", page_title="PragyanAI Job Portal")

    # --- Session State Initialization ---
    if 'page' not in st.session_state: st.session_state.page = "login"
    
    # AI features state
    if 'parsed' not in st.session_state: st.session_state.parsed = {}
    if 'full_text' not in st.session_state: st.session_state.full_text = ""
    if 'excel_data' not in st.session_state: st.session_state.excel_data = None
    if 'qa_answer' not in st.session_state: st.session_state.qa_answer = ""
    if 'iq_output' not in st.session_state: st.session_state.iq_output = ""
    if 'jd_fit_output' not in st.session_state: st.session_state.jd_fit_output = ""
    if 'interview_qa' not in st.session_state: st.session_state.interview_qa = [] 
    if 'evaluation_report' not in st.session_state: st.session_state.evaluation_report = ""
    
    # Admin Dashboard specific lists
    if 'admin_jd_list' not in st.session_state: st.session_state.admin_jd_list = [] 
    if 'resumes_to_analyze' not in st.session_state: st.session_state.resumes_to_analyze = [] 
    if 'admin_match_results' not in st.session_state: st.session_state.admin_match_results = [] 
    if 'resume_statuses' not in st.session_state: st.session_state.resume_statuses = {} 
        
    # Vendor State Init
    if 'vendors' not in st.session_state: st.session_state.vendors = []
    if 'vendor_statuses' not in st.session_state: st.session_state.vendor_statuses = {}
        
    # Candidate Dashboard specific lists
    if 'candidate_jd_list' not in st.session_state: st.session_state.candidate_jd_list = []
    if 'candidate_match_results' not in st.session_state: st.session_state.candidate_match_results = []
    
    # Resume Parsing Upload State
    if 'candidate_uploaded_resumes' not in st.session_state: st.session_state.candidate_uploaded_resumes = []
    if 'pasted_cv_text' not in st.session_state: st.session_state.pasted_cv_text = "" 
    
    # CV Builder Form State
    if "cv_form_data" not in st.session_state: 
        st.session_state.cv_form_data = {
            "name": "", "email": "", "phone": "", "linkedin": "", "github": "",
            "skills": [], "experience": [], "education": [], "certifications": [], 
            "projects": [], "strength": [], "personal_details": ""
        }
    
    # Filter State
    if "candidate_filter_skills_multiselect" not in st.session_state:
        st.session_state.candidate_filter_skills_multiselect = []
    if "filtered_jds_display" not in st.session_state:
        st.session_state.filtered_jds_display = []
    if "last_selected_skills" not in st.session_state:
        st.session_state.last_selected_skills = []


    # --- Page Routing ---
    if st.session_state.page == "login":
        login_page()
    elif st.session_state.page == "signup":
        signup_page()
    elif st.session_state.page == "admin_dashboard":
        # Pass necessary utility functions to the dashboard
        admin_dashboard(go_to, extract_jd_metadata, parse_and_store_resume, evaluate_jd_fit, get_file_type, extract_content, extract_jd_from_linkedin_url)
    elif st.session_state.page == "candidate_dashboard":
        candidate_dashboard(go_to, parse_and_store_resume, qa_on_resume, evaluate_interview_answers, generate_interview_questions, question_section_options, extract_jd_metadata, get_file_type, extract_content, extract_jd_from_linkedin_url, clear_interview_state, evaluate_jd_fit, DEFAULT_JOB_TYPES, DEFAULT_ROLES)
    elif st.session_state.page == "hiring_dashboard":
        hiring_dashboard(go_to)

if __name__ == '__main__':
    # Create the necessary dashboard files if they don't exist
    for filename in ['admin_dashboard.py', 'candidate_dashboard.py', 'hiring_dashboard.py']:
        if not os.path.exists(filename):
            print(f"WARNING: The file '{filename}' is missing. Please create it using the provided code blocks.")
            
    main()
