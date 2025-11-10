# =================================================================
# app_utils.py
# Contains all common configurations, core logic, and utility functions
# =================================================================

import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
import traceback
import re
from dotenv import load_dotenv 
from datetime import date 
import csv 
from streamlit.runtime.uploaded_file_manager import UploadedFile


# -------------------------
# CONFIGURATION & API SETUP
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"

# Options for LLM functions
section_options = ["name", "email", "phone", "skills", "education", "experience", "certifications", "projects", "strength", "personal_details", "github", "linkedin", "full resume"]
question_section_options = ["skills","experience", "certifications", "projects", "education"] 

# Default Categories for JD Filtering (New)
DEFAULT_JOB_TYPES = ["Full-time", "Contract", "Internship", "Remote", "Part-time"]
DEFAULT_ROLES = ["Software Engineer", "Data Scientist", "Product Manager", "HR Manager", "Marketing Specialist", "Operations Analyst"]

# Load environment variables from .env file
load_dotenv()

# Ensure GROQ_API_KEY is defined
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

if not GROQ_API_KEY:
    st.warning(
        "🚨 WARNING: GROQ_API_KEY environment variable not set. "
        "AI functionality (Parsing, Matching, Q&A) will not work. "
        "Please ensure a '.env' file exists with your key."
    )
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set. AI functions disabled.")
            return Completions()
    
    client = MockGroqClient()
else:
    # Initialize Groq Client
    client = Groq(api_key=GROQ_API_KEY)

# -------------------------
# Utility: Navigation Manager
# -------------------------
def go_to(page_name):
    """Changes the current page in Streamlit's session state."""
    st.session_state.page = page_name

def clear_interview_state():
    """Clears all generated questions, answers, and the evaluation report."""
    st.session_state.interview_qa = []
    st.session_state.iq_output = ""
    st.session_state.evaluation_report = ""
    st.toast("Practice answers cleared.")
    
# -------------------------
# CORE LOGIC: FILE HANDLING AND EXTRACTION
# -------------------------

def get_file_type(file_path):
    """Identifies the file type based on its extension."""
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    if ext == 'pdf':
        return 'pdf'
    elif ext == 'docx':
        return 'docx'
    elif ext == 'xlsx':
        return 'xlsx'
    elif ext in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf']:
        return ext
    else:
        return 'txt' 

def extract_content(file_type, file_path):
    """Extracts text content from various file types."""
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
        
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                text += f"--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        text += row_text + '\n'
                text += "\n"
        
        elif file_type == 'csv':
             with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += ' | '.join(row) + '\n'
        
        elif file_type in ['txt', 'json', 'md', 'markdown', 'csv', 'rtf'] or file_type not in ['pdf', 'docx', 'xlsx']:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

        if not text.strip():
            return f"Error: {file_type.upper()} content extraction failed. The file appears empty or non-text content could not be read."
        
        return text
    
    except Exception as e:
        return f"Fatal Extraction Error: Failed to read file content ({file_type}). Error details: {e}"

# -------------------------
# LLM & Extraction Functions
# -------------------------

@st.cache_data(show_spinner="Extracting JD metadata...")
def extract_jd_metadata(jd_text):
    """Extracts structured metadata (Role, Job Type, Key Skills) from raw JD text."""
    if not GROQ_API_KEY:
        return {"role": "N/A", "job_type": "N/A", "key_skills": []}

    prompt = f"""Analyze the following Job Description and extract the key metadata.
    Job Description: {jd_text}
    Provide the output strictly as a JSON object with the following three keys:
    1.  **role**: The main job title (e.g., 'Data Scientist', 'Senior Software Engineer'). If not clear, default to 'General Analyst'.
    2.  **job_type**: The employment type (e.g., 'Full-time', 'Contract', 'Internship', 'Remote'). If not clear, default to 'Full-time'.
    3.  **key_skills**: A list of 5 to 10 most critical hard and soft skills required (e.g., ['Python', 'AWS', 'Teamwork', 'SQL']).
    """
    content = ""
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        content = response.choices[0].message.content.strip()

        json_str = content
        if json_str.startswith('```json'): json_str = json_str[len('```json'):]
        if json_str.endswith('```'): json_str = json_str[:-len('```')]
        json_str = json_str.strip()

        json_start = json_str.find('{')
        json_end = json_str.rfind('}') + 1

        if json_start != -1 and json_end != -1 and json_end > json_start:
            json_str = json_str[json_start:json_end]

        parsed = json.loads(json_str)
        
        return {
            "role": parsed.get("role", "General Analyst"),
            "job_type": parsed.get("job_type", "Full-time"),
            "key_skills": [s.strip() for s in parsed.get("key_skills", []) if isinstance(s, str)]
        }

    except Exception:
        return {"role": "General Analyst (LLM Error)", "job_type": "Full-time (LLM Error)", "key_skills": ["LLM Error", "Fallback"]}


@st.cache_data(show_spinner="Analyzing content with Groq LLM...")
def parse_with_llm(text, return_type='json'):
    """Sends resume text to the LLM for structured information extraction."""
    if text.startswith("Error"):
        return {"error": text, "raw_output": ""}
    if not GROQ_API_KEY:
        return {"error": "GROQ_API_KEY not set. Cannot run LLM parsing.", "raw_output": ""}

    prompt = f"""Extract the following information from the resume in structured JSON.
    Ensure all relevant details for each category are captured.
    - Name, - Email, - Phone, - Skills, - Education, - Experience, - Certifications, 
    - Projects, - Strength, - Personal Details, - Github (URL), - LinkedIn (URL)
    Resume Text: {text}
    Provide the output strictly as a JSON object.
    """
    content = ""
    parsed = {}
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
        content = response.choices[0].message.content.strip()

        json_str = content
        if json_str.startswith('```json'): json_str = json_str[len('```json'):]
        if json_str.endswith('```'): json_str = json_str[:-len('```')]
        json_str = json_str.strip()

        json_start = json_str.find('{')
        json_end = json_str.rfind('}') + 1

        if json_start != -1 and json_end != -1 and json_end > json_start:
            json_str = json_str[json_start:json_end]
        else:
            raise json.JSONDecodeError("Could not isolate a valid JSON structure from LLM response.", content, 0)

        parsed = json.loads(json_str)

    except (json.JSONDecodeError, ValueError) as e:
        error_msg = f"JSON decoding or LLM API error. Error: {e}"
        parsed = {"error": error_msg, "raw_output": content}
    except Exception as e:
        error_msg = f"LLM API interaction error: {e}"
        parsed = {"error": error_msg, "raw_output": "No LLM response due to API error."}

    if return_type == 'json':
        return parsed
    
    # ... (Markdown formatting logic can stay here if needed, but simplified for brevity)
    return parsed

def extract_jd_from_linkedin_url(url: str) -> str:
    """Simulates JD content extraction from a LinkedIn URL."""
    try:
        job_title = "Data Scientist"
        match = re.search(r'/jobs/view/([^/]+)', url) or re.search(r'/jobs/(\w+)', url)
        if match:
            job_title = match.group(1).split('?')[0].replace('-', ' ').title()
        
        if "linkedin.com/jobs/" not in url:
             return f"[Error: Not a valid LinkedIn Job URL format: {url}]"

        jd_text = f"""
        --- Simulated JD for: {job_title} ---
        **Company:** Quantum Analytics Inc.
        **Role:** {job_title}
        **Responsibilities:**
        - Develop and implement machine learning models.
        - Clean, transform, and analyze large datasets using Python/R and SQL.
        **Requirements:**
        - MS/PhD in Computer Science.
        - 3+ years of experience as a Data Scientist.
        - Expertise in Python (Pandas, Scikit-learn).
        - Experience with cloud platforms (AWS, Azure, or GCP).
        --- End Simulated JD ---
        """
        return jd_text.strip()
            
    except Exception as e:
        return f"[Fatal Extraction Error: Simulation failed for URL {url}. Error: {e}]"


def evaluate_jd_fit(job_description, parsed_json):
    """Evaluates how well a resume fits a given job description, including section-wise scores."""
    if not GROQ_API_KEY: return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if not job_description.strip(): return "Please paste a job description."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."
    
    relevant_resume_data = {
        'Skills': parsed_json.get('skills', 'Not found or empty'),
        'Experience': parsed_json.get('experience', 'Not found or empty'),
        'Education': parsed_json.get('education', 'Not found or empty'),
    }
    resume_summary = json.dumps(relevant_resume_data, indent=2)

    prompt = f"""Evaluate how well the following resume content matches the provided job description.
    Job Description: {job_description}
    Resume Sections for Analysis: {resume_summary}
    Provide a detailed evaluation structured as follows:
    1.  **Overall Fit Score:** A score out of 10.
    2.  **Section Match Percentages:** A percentage score for the match in the key sections (Skills, Experience, Education).
    3.  **Strengths/Matches:** Key points where the resume aligns well with the JD.
    4.  **Gaps/Areas for Improvement:** Key requirements in the JD that are missing or weak in the resume.
    5.  **Overall Summary:** A concise summary of the fit.
    
    Format the output strictly as follows:
    Overall Fit Score: [Score]/10
    
    --- Section Match Analysis ---
    Skills Match: [XX]%
    Experience Match: [YY]%
    Education Match: [ZZ]%
    
    Strengths/Matches:
    - Point 1
    - Point 2
    
    Gaps/Areas for Improvement:
    - Point 1
    - Point 2
    
    Overall Summary: [Concise summary]
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def evaluate_interview_answers(qa_list, parsed_json):
    """Evaluates the user's answers against the resume content and provides feedback."""
    if not GROQ_API_KEY: return "AI Evaluation Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot evaluate due to resume parsing errors."

    resume_summary = json.dumps(parsed_json, indent=2)
    qa_summary = "\n---\n".join([f"Q: {item['question']}\nA: {item['answer']}" for item in qa_list])
    total_questions = len(qa_list)
    
    prompt = f"""You are an expert HR Interviewer. Evaluate the candidate's answers based on the following:
    1.  **The Candidate's Resume Content (for context):** {resume_summary}
    2.  **The Candidate's Questions and Answers:** {qa_summary}
    
    For each Question-Answer pair, provide a score (out of 10) and detailed feedback. The feedback must include:
    * Clarity & Accuracy: How well the answer directly and accurately addresses the question.
    * Gaps & Improvements: Specific suggestions for improvement.
    
    Finally, provide an **Overall Summary** and a **Total Score (out of {total_questions * 10})**.
    
    Format the output strictly using Markdown headings and bullet points.
    """

    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.3
    )
    return response.choices[0].message.content.strip()


def generate_interview_questions(parsed_json, section):
    """Generates categorized interview questions using LLM."""
    if not GROQ_API_KEY: return "AI Functions Disabled: GROQ_API_KEY not set."
    if "error" in parsed_json: return "Cannot generate questions due to resume parsing errors."
    
    section_title = section.replace("_", " ").title()
    section_content = parsed_json.get(section, "")
    if isinstance(section_content, (list, dict)):
        section_content = json.dumps(section_content, indent=2)

    if not section_content.strip():
        return f"No significant content found for the '{section_title}' section."

    prompt = f"""Based on the following {section_title} section from the resume: {section_content}
Generate 3 interview questions each for these levels: Generic, Basic, Intermediate, Difficult.
IMPORTANT: Format the output strictly as follows, with level headers and questions starting with 'Qx:':
[Generic]
Q1: Question text...
...
"""
    response = client.chat.completions.create(
        model=GROQ_MODEL, 
        messages=[{"role": "user", "content": prompt}], 
        temperature=0.5
    )
    return response.choices[0].message.content.strip()

def dump_to_excel(parsed_json, filename):
    """Dumps parsed JSON data to an Excel file."""
    # (Existing logic for dump_to_excel)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Data"
    ws.append(["Category", "Details"])
    
    section_order = ['name', 'email', 'phone', 'github', 'linkedin', 'experience', 'education', 'skills', 'projects', 'certifications', 'strength', 'personal_details']
    
    for section_key in section_order:
        if section_key in parsed_json and parsed_json[section_key]:
            content = parsed_json[section_key]
            
            if section_key in ['name', 'email', 'phone', 'github', 'linkedin']:
                ws.append([section_key.replace('_', ' ').title(), str(content)])
            else:
                ws.append([])
                ws.append([section_key.replace('_', ' ').title()])
                
                if isinstance(content, list):
                    for item in content:
                        if item:
                            ws.append(["", str(item)])
                elif isinstance(content, dict):
                    for k, v in content.items():
                        if v:
                            ws.append(["", f"{k.replace('_', ' ').title()}: {v}"])
                else:
                    ws.append(["", str(content)])

    wb.save(filename)
    with open(filename, "rb") as f:
        return f.read()

def parse_and_store_resume(file_input, file_name_key='default', source_type='file'):
    """Handles file/text input, parsing, and stores results."""
    text = None
    file_name = f"Pasted Text ({date.today().strftime('%Y-%m-%d')})"

    if source_type == 'file':
        if not isinstance(file_input, UploadedFile):
            return {"error": "Invalid file input type passed to parser.", "full_text": ""}

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file_input.name) 
        with open(temp_path, "wb") as f:
            f.write(file_input.getbuffer()) 

        file_type = get_file_type(temp_path)
        text = extract_content(file_type, temp_path)
        file_name = file_input.name.split('.')[0]
    
    elif source_type == 'text':
        text = file_input
        
    if text.startswith("Error"):
        return {"error": text, "full_text": text, "name": file_name}

    parsed = parse_with_llm(text, return_type='json')
    
    if not parsed or "error" in parsed:
        return {"error": parsed.get('error', 'Unknown parsing error'), "full_text": text, "name": file_name}

    excel_data = None
    if file_name_key == 'single_resume_candidate':
        try:
            name = parsed.get('name', 'candidate').replace(' ', '_').strip()
            name = "".join(c for c in name if c.isalnum() or c in ('_', '-')).rstrip()
            if not name: name = "candidate"
            excel_filename = os.path.join(tempfile.gettempdir(), f"{name}_parsed_data.xlsx")
            excel_data = dump_to_excel(parsed, excel_filename)
        except Exception:
            pass
    
    final_name = parsed.get('name', file_name)

    return {
        "parsed": parsed,
        "full_text": text,
        "excel_data": excel_data,
        "name": final_name
    }


def qa_on_resume(question):
    """Chatbot for Resume (Q&A) using LLM."""
    if not GROQ_API_KEY:
        return "AI Chatbot Disabled: GROQ_API_KEY not set."
        
    parsed_json = st.session_state.parsed
    full_text = st.session_state.full_text
    prompt = f"""Given the following resume information:
    Resume Text: {full_text}
    Parsed Resume Data (JSON): {json.dumps(parsed_json, indent=2)}
    Answer the following question about the resume concisely and directly.
    If the information is not present, state that clearly.
    Question: {question}
    """
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.4)
    return response.choices[0].message.content.strip()

def generate_cv_html(parsed_data):
    """Generates a simple, print-friendly HTML string from parsed data for PDF conversion."""
    
    css = """
    <style>
        @page { size: A4; margin: 1cm; }
        body { font-family: 'Arial', sans-serif; line-height: 1.5; margin: 0; padding: 0; font-size: 10pt; }
        .header { text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 20px; }
        .header h1 { margin: 0; font-size: 1.8em; }
        .contact-info { display: flex; justify-content: center; font-size: 0.8em; color: #555; }
        .contact-info span { margin: 0 8px; }
        .section { margin-bottom: 15px; page-break-inside: avoid; }
        .section h2 { border-bottom: 1px solid #999; padding-bottom: 3px; margin-bottom: 8px; font-size: 1.1em; text-transform: uppercase; color: #333; }
        .item-list ul { list-style-type: disc; margin-left: 20px; padding-left: 0; margin-top: 0; }
        .item-list ul li { margin-bottom: 3px; }
        .item-list p { margin: 3px 0 8px 0; }
        a { color: #0056b3; text-decoration: none; }
    </style>
    """
    
    html_content = f"<html><head>{css}<title>{parsed_data.get('name', 'CV')}</title></head><body>"
    
    # 1. Header and Contact Info
    html_content += '<div class="header">'
    html_content += f"<h1>{parsed_data.get('name', 'Candidate Name')}</h1>"
    
    contact_parts = []
    if parsed_data.get('email'): contact_parts.append(f"<span>📧 {parsed_data['email']}</span>")
    if parsed_data.get('phone'): contact_parts.append(f"<span>📱 {parsed_data['phone']}</span>")
    if parsed_data.get('linkedin'): contact_parts.append(f"<span>🔗 <a href='{parsed_data['linkedin']}'>LinkedIn</a></span>")
    if parsed_data.get('github'): contact_parts.append(f"<span>💻 <a href='{parsed_data['github']}'>GitHub</a></span>")
    
    html_content += f'<div class="contact-info">{" | ".join(contact_parts)}</div>'
    html_content += '</div>'
    
    # 2. Sections
    section_order = ['personal_details', 'experience', 'projects', 'education', 'certifications', 'skills', 'strength']
    
    for k in section_order:
        v = parsed_data.get(k)
        
        if k in ['name', 'email', 'phone', 'linkedin', 'github']: continue 

        if v and (isinstance(v, str) and v.strip() or isinstance(v, list) and v):
            
            html_content += f'<div class="section"><h2>{k.replace("_", " ").title()}</h2>'
            html_content += '<div class="item-list">'
            
            if k == 'personal_details' and isinstance(v, str):
                html_content += f"<p>{v}</p>"
            elif isinstance(v, list):
                html_content += '<ul>'
                for item in v:
                    if item: 
                        html_content += f"<li>{item}</li>"
                html_content += '</ul>'
            else:
                html_content += f"<p>{v}</p>"
                
            html_content += '</div></div>'

    html_content += '</body></html>'
    return html_content

def format_parsed_json_to_markdown(parsed_data):
    """Formats the parsed JSON data into a clean, CV-like Markdown structure."""
    md = ""
    
    if parsed_data.get('name'):
        md += f"# **{parsed_data['name']}**\n\n"
    
    contact_info = []
    if parsed_data.get('email'): contact_info.append(parsed_data['email'])
    if parsed_data.get('phone'): contact_info.append(parsed_data['phone'])
    if parsed_data.get('linkedin'): contact_info.append(f"[LinkedIn]({parsed_data['linkedin']})")
    if parsed_data.get('github'): contact_info.append(f"[GitHub]({parsed_data['github']})")
    
    if contact_info:
        md += f"| {' | '.join(contact_info)} |\n"
        md += "| " + " | ".join(["---"] * len(contact_info)) + " |\n\n"
    
    section_order = ['personal_details', 'experience', 'projects', 'education', 'certifications', 'skills', 'strength']
    
    for k in section_order:
        v = parsed_data.get(k)
        
        if k in ['name', 'email', 'phone', 'linkedin', 'github']: continue 

        if v and (isinstance(v, str) and v.strip() or isinstance(v, list) and v):
            
            md += f"## **{k.replace('_', ' ').upper()}**\n"
            md += "---\n"
            
            if k == 'personal_details' and isinstance(v, str):
                md += f"{v}\n\n"
            elif isinstance(v, list):
                for item in v:
                    if item: 
                        md += f"- {item}\n"
                md += "\n"
            else:
                md += f"{v}\n\n"
    return md
